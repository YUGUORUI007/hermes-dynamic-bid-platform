from __future__ import annotations

import csv
import re
from datetime import datetime
from pathlib import Path


HEADER_FIELD_MAP = {
    "序号": "sequence",
    "项目名称": "name",
    "投标性质": "bid_mode_raw",
    "是否已报名": "signup_state_raw",
    "报名截止时间": "signup_deadline_raw",
    "投标时间(开标时间)": "bid_datetime_raw",
    "投标地点": "bid_location",
    "保证金金额(万元)": "deposit_amount_raw",
    "保证金截止时间": "deposit_deadline_raw",
    "保证金是否已交": "deposit_state_raw",
    "招标编号": "tender_code",
    "联系人": "contact_name",
    "联系电话": "contact_phone",
    "合同期限": "contract_term",
    "备注": "notes_raw",
}


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", "\n")
    text = re.sub(r"\s*\n\s*", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_header(value: object) -> str:
    text = clean_cell(value)
    text = text.replace("（", "(").replace("）", ")")
    text = text.replace("：", ":")
    text = re.sub(r"\s+", "", text)
    return text


def parse_amount(value: str) -> str:
    text = clean_cell(value)
    if not text:
        return ""
    match = re.search(r"\d+(?:\.\d+)?", text)
    return match.group(0) if match else ""


def parse_datetime_text(value: str) -> str:
    text = clean_cell(value)
    if not text:
        return ""

    text = (
        text.replace("年", "-")
        .replace("月", "-")
        .replace("日", " ")
        .replace("：", ":")
        .replace("（", " ")
        .replace("）", " ")
    )
    is_pm = "下午" in text
    is_am = "上午" in text
    text = (
        text.replace("上午", " ")
        .replace("下午", " ")
        .replace("前到账", " ")
        .replace("到账", " ")
        .replace("前", " ")
        .replace("止", " ")
    )
    text = re.sub(r"\s+", " ", text).strip()

    date_match = re.search(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", text)
    if not date_match:
        return ""

    year, month, day = [int(item) for item in date_match.groups()]
    hour = 0
    minute = 0
    time_match = re.search(r"(\d{1,2}):(\d{1,2})", text)
    if time_match:
        hour, minute = [int(item) for item in time_match.groups()]
        if is_pm and hour < 12:
            hour += 12
        if is_am and hour == 12:
            hour = 0

    try:
        return datetime(year, month, day, hour, minute).strftime("%Y-%m-%d %H:%M")
    except ValueError:
        return ""


def parse_submission_datetime(note_text: str) -> str:
    text = clean_cell(note_text)
    if not text:
        return ""
    match = re.search(r"文件递交\s*(\d{4}-\d{1,2}-\d{1,2})\s*(\d{1,2}:\d{1,2})", text)
    if not match:
        return ""
    return f"{match.group(1)} {match.group(2)}"


def parse_bid_mode(value: str) -> str:
    text = clean_cell(value)
    if "陪标" in text:
        return "partner"
    return "self"


def infer_project_type(name: str, original_name: str) -> str:
    combined = f"{name} {original_name}"
    if "物业" in combined:
        return "物业服务"
    return ""


def infer_status(bid_datetime: str) -> str:
    if not bid_datetime:
        return "tracking"
    try:
        bid_at = datetime.strptime(bid_datetime, "%Y-%m-%d %H:%M")
    except ValueError:
        return "tracking"
    return "result_pending" if bid_at < datetime.now() else "tracking"


def extract_agency(note_text: str) -> str:
    text = clean_cell(note_text)
    if not text:
        return ""
    match = re.search(r"(?:代理|代理机构)[:：]\s*([^；;]+)", text)
    return match.group(1).strip() if match else ""


def extract_file_fee(note_text: str) -> str:
    text = clean_cell(note_text)
    if not text:
        return ""
    for part in re.split(r"[；;]", text):
        if "文件" in part and "元" in part:
            return part.strip()
    return ""


def extract_payment_info(note_text: str) -> str:
    text = clean_cell(note_text)
    if not text:
        return ""
    parts = []
    for part in re.split(r"[；;]", text):
        if any(keyword in part for keyword in ("收款", "开户", "银行", "账号")):
            parts.append(part.strip())
    return "；".join(parts)


def build_notes(row: dict[str, str]) -> str:
    parts: list[str] = []
    signup_state = clean_cell(row.get("signup_state_raw"))
    deposit_state = clean_cell(row.get("deposit_state_raw"))
    raw_notes = clean_cell(row.get("notes_raw"))
    if signup_state:
        parts.append(f"报名状态：{signup_state}")
    if deposit_state:
        parts.append(f"保证金状态：{deposit_state}")
    if raw_notes:
        parts.append(raw_notes)
    return "\n".join(parts)


def map_row_to_project_form(row: dict[str, str], original_name: str) -> dict[str, str]:
    name = clean_cell(row.get("name"))
    signup_deadline_raw = clean_cell(row.get("signup_deadline_raw"))
    signup_deadline = parse_datetime_text(signup_deadline_raw)
    document_sale_deadline = ""
    if "文件发售" in signup_deadline_raw or "发售截止" in signup_deadline_raw:
        document_sale_deadline = signup_deadline
        signup_deadline = ""

    bid_datetime = parse_datetime_text(clean_cell(row.get("bid_datetime_raw")))
    deposit_deadline = parse_datetime_text(clean_cell(row.get("deposit_deadline_raw")))
    notes_raw = clean_cell(row.get("notes_raw"))
    agency = extract_agency(notes_raw)
    file_fee = extract_file_fee(notes_raw)
    payment_info = extract_payment_info(notes_raw)

    return {
        "name": name,
        "short_name": name[:18],
        "tender_code": clean_cell(row.get("tender_code")),
        "buyer": "",
        "project_type": infer_project_type(name, original_name),
        "bid_mode": parse_bid_mode(clean_cell(row.get("bid_mode_raw"))),
        "status": infer_status(bid_datetime),
        "owner_name": "",
        "agency": agency,
        "contact_name": clean_cell(row.get("contact_name")),
        "contact_phone": clean_cell(row.get("contact_phone")),
        "location": clean_cell(row.get("bid_location")),
        "service_scope": "",
        "contract_term": clean_cell(row.get("contract_term")),
        "budget_amount": "",
        "deposit_amount": parse_amount(clean_cell(row.get("deposit_amount_raw"))),
        "signup_deadline": signup_deadline,
        "document_sale_deadline": document_sale_deadline,
        "clarification_deadline": "",
        "site_visit_time": "",
        "deposit_deadline": deposit_deadline,
        "bid_datetime": bid_datetime,
        "submission_datetime": parse_submission_datetime(notes_raw),
        "bid_location": clean_cell(row.get("bid_location")),
        "file_fee": file_fee,
        "payment_info": payment_info,
        "notes": build_notes(row),
    }


def _read_xlsx_rows(file_path: Path) -> list[list[object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("当前环境未安装 openpyxl，暂时无法导入 Excel 台账。") from exc

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook.active
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def _read_csv_rows(file_path: Path) -> list[list[object]]:
    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "gb18030", "utf-8"):
        try:
            with file_path.open("r", encoding=encoding, newline="") as handle:
                return [list(row) for row in csv.reader(handle)]
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ValueError(f"CSV 文件编码无法识别：{last_error}") from last_error


def read_ledger_rows(file_path: Path, original_name: str) -> list[dict[str, str]]:
    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        raw_rows = _read_xlsx_rows(file_path)
    elif suffix == ".csv":
        raw_rows = _read_csv_rows(file_path)
    else:
        raise ValueError("当前仅支持导入 .xlsx 或 .csv 台账。")

    header_index = None
    mapped_headers: list[str | None] = []
    for index, row in enumerate(raw_rows):
        normalized = [normalize_header(cell) for cell in row]
        if "项目名称" not in normalized:
            continue
        header_index = index
        mapped_headers = [HEADER_FIELD_MAP.get(item) for item in normalized]
        break

    if header_index is None:
        raise ValueError("未在台账中识别到表头，请确认文件包含“项目名称”等列。")

    rows: list[dict[str, str]] = []
    for row in raw_rows[header_index + 1 :]:
        if not any(clean_cell(cell) for cell in row):
            continue
        row_dict: dict[str, str] = {}
        for cell_index, field_name in enumerate(mapped_headers):
            if not field_name:
                continue
            value = row[cell_index] if cell_index < len(row) else ""
            row_dict[field_name] = clean_cell(value)

        project_name = clean_cell(row_dict.get("name"))
        if not project_name or project_name.startswith("★"):
            continue
        rows.append(map_row_to_project_form(row_dict, original_name))
    return rows
